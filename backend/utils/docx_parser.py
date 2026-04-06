"""
DOCX field extractor and filler.
Finds yellow-highlighted runs → extracts placeholder text + context.
On export: replaces highlighted runs with filled values.

═══ IMPROVEMENTS v2 ═══
1. 8 field types: date, time_range, number, percentage, name, short, sentence, paragraph, bullet_list
2. Bullet/list detection from placeholder structure (-, •, a), b), 1., 2.)
3. Smarter Vietnamese banking patterns (nghị quyết, văn bản, đơn vị, chức danh)
4. Rich type hints with format examples and negative examples
5. Context-aware classification: xem cả paragraph trước/sau để hiểu loại trường
"""
import os
import re
import re as _re
import subprocess
import tempfile
import zipfile
import xml.etree.ElementTree as ET
from typing import List, Dict

NS_W = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
ET.register_namespace('wpc', 'http://schemas.microsoft.com/office/word/2010/wordprocessingCanvas')
ET.register_namespace('cx',  'http://schemas.microsoft.com/office/drawing/2014/chartex')
ET.register_namespace('mc',  'http://schemas.openxmlformats.org/markup-compatibility/2006')
ET.register_namespace('o',   'urn:schemas-microsoft-com:office:office')
ET.register_namespace('r',   'http://schemas.openxmlformats.org/officeDocument/2006/relationships')
ET.register_namespace('m',   'http://schemas.openxmlformats.org/officeDocument/2006/math')
ET.register_namespace('v',   'urn:schemas-microsoft-com:vml')
ET.register_namespace('wp14','http://schemas.microsoft.com/office/word/2010/wordprocessingDrawing')
ET.register_namespace('wp',  'http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing')
ET.register_namespace('w10', 'urn:schemas-microsoft-com:office:word')
ET.register_namespace('w',   'http://schemas.openxmlformats.org/wordprocessingml/2006/main')
ET.register_namespace('w14', 'http://schemas.microsoft.com/office/word/2010/wordml')
ET.register_namespace('w15', 'http://schemas.microsoft.com/office/word/2012/wordml')
ET.register_namespace('w16cex','http://schemas.microsoft.com/office/word/2018/wordml/cex')
ET.register_namespace('w16se','http://schemas.microsoft.com/office/word/2015/wordml/symex')
ET.register_namespace('wpg', 'http://schemas.microsoft.com/office/word/2010/wordprocessingGroup')
ET.register_namespace('wpi', 'http://schemas.microsoft.com/office/word/2010/wordprocessingInk')
ET.register_namespace('wne', 'http://schemas.microsoft.com/office/word/2006/wordml')
ET.register_namespace('wps', 'http://schemas.microsoft.com/office/word/2010/wordprocessingShape')
ET.register_namespace('xml', 'http://www.w3.org/XML/1998/namespace')


# ─── Extract yellow fields ────────────────────────────────────────────────────

def extract_fields(docx_bytes: bytes) -> List[Dict]:
    with tempfile.TemporaryDirectory() as tmp:
        docx_path  = os.path.join(tmp, "template.docx")
        unpack_dir = os.path.join(tmp, "unpacked")
        with open(docx_path, "wb") as f:
            f.write(docx_bytes)
        with zipfile.ZipFile(docx_path, "r") as z:
            z.extractall(unpack_dir)
        return _parse_highlights(os.path.join(unpack_dir, "word", "document.xml"))


# ═══════════════════════════════════════════════════════════════════════════════
# IMPROVED: Field Type Detection Patterns
# ═══════════════════════════════════════════════════════════════════════════════

# ── Date patterns ─────────────────────────────────────────────────────────────
_DATE_PATTERNS = [
    _re.compile(r'\bDD/MM/YYYY\b', _re.I),
    _re.compile(r'\b\d{1,2}/\d{1,2}/\d{2,4}\b'),
    _re.compile(r'\bngày\s+\d{1,2}\b', _re.I),
    _re.compile(r'\btháng\s*\d{1,2}\b', _re.I),
    _re.compile(r'\bnăm\s+20\d{2}\b', _re.I),
    _re.compile(r'\b20\d{2}\b'),
    _re.compile(r'ngày\s*[.…]{2,}|tháng\s*[.…]{2,}|năm\s*[.…]{2,}', _re.I),
    _re.compile(r'[.…]{3,}\s*/\s*[.…]{3,}\s*/\s*[.…]{3,}'),  # .../.../.../
]

# ── Time range / period patterns ──────────────────────────────────────────────
_TIME_RANGE_PATTERNS = [
    _re.compile(r'\bQuý\s+[IVX\d]+[/\s]*\d{4}\b', _re.I),
    _re.compile(r'\bgiai\s*đoạn\b', _re.I),
    _re.compile(r'\btừ\s+(tháng|năm|ngày)\b.*\bđến\b', _re.I),
    _re.compile(r'\bnăm\s+\d{4}\s*[-–]\s*\d{4}\b'),
    _re.compile(r'\b\d{1,2}\s*tháng\s*(đầu|cuối)\s*năm\b', _re.I),
    _re.compile(r'\b(nửa đầu|nửa cuối|đầu năm|cuối năm|cả năm)\b', _re.I),
    _re.compile(r'\bkỳ\s+báo\s*cáo\b', _re.I),
    _re.compile(r'\bthời\s*(gian|kỳ|điểm)\b', _re.I),
]

# ── Number / quantity patterns ────────────────────────────────────────────────
_NUMBER_PATTERNS = [
    _re.compile(r'\d+[.,]\d+\s*(tỷ|triệu|nghìn|người|dự án|lệnh|đồng|KHCN|KH)', _re.I),
    _re.compile(r'\bsố\s*(lượng|liệu|tiền|vốn|dự án|giao dịch|lệnh|người)\b', _re.I),
    _re.compile(r'\[\s*số\s*\]|\[\s*tỷ\s*\]', _re.I),
    _re.compile(r'\bXXX\s*(tỷ|triệu|nghìn|đồng|%)', _re.I),
    _re.compile(r'\b\d+\s*(tỷ|triệu|nghìn|đồng)\b', _re.I),
]

# ── Percentage patterns ───────────────────────────────────────────────────────
_PERCENTAGE_PATTERNS = [
    _re.compile(r'\d+[.,]?\d*\s*%'),
    _re.compile(r'\[\s*%\s*\]'),
    _re.compile(r'\bXXX\s*%'),
    _re.compile(r'\btỷ\s*lệ\b', _re.I),
    _re.compile(r'\b(đạt|tăng|giảm)\s+\d', _re.I),
]

# ── Name / title / entity patterns ────────────────────────────────────────────
_NAME_PATTERNS = [
    _re.compile(r'\b(Ông|Bà|Anh|Chị|Đ/c|đồng chí)\s', _re.I),
    _re.compile(r'\bchức\s*(danh|vụ)\b', _re.I),
    _re.compile(r'\b(Giám đốc|Phó Giám đốc|Trưởng phòng|Phó phòng|Chủ tịch|Tổng Giám đốc)\b', _re.I),
    _re.compile(r'\b(Nghị quyết|Quyết định|Thông tư|Nghị định|Chỉ thị|Công văn)\s*số\b', _re.I),
    _re.compile(r'\bsố\s+\d+[/\-]', _re.I),
    _re.compile(r'\b(NQ|QĐ|TT|NĐ|CT|CV)\s*\d+', _re.I),
    _re.compile(r'\b[A-ZÀ-Ỹ]{2,}\b'),  # viết hoa >= 2 ký tự liên tục = tên viết tắt
]

# ── Bullet list patterns ─────────────────────────────────────────────────────
_BULLET_PATTERNS = [
    _re.compile(r'^\s*[-–—•]\s', _re.M),           # - hoặc • đầu dòng
    _re.compile(r'^\s*[a-zđ]\)\s', _re.M),          # a) b) c)
    _re.compile(r'^\s*\d+[.)]\s', _re.M),            # 1. 2. 3) 4)
    _re.compile(r'^\s*\([a-zđivx\d]+\)\s', _re.M),   # (i) (ii) (a) (1)
    _re.compile(r';\s*\n', _re.M),                    # dấu ; xuống dòng = list
    _re.compile(r'(Thứ nhất|Thứ hai|Thứ ba|Một là|Hai là)', _re.I),
]


# ═══════════════════════════════════════════════════════════════════════════════
# IMPROVED: Type Hints — detailed, with format examples + negative examples
# ═══════════════════════════════════════════════════════════════════════════════

_FIELD_TYPE_HINTS = {
    "date": (
        "ĐÂY LÀ TRƯỜNG NGÀY THÁNG CỤ THỂ.\n"
        "→ Điền MỘT ngày/tháng/năm duy nhất từ tài liệu nguồn hoặc ngày hiện tại.\n"
        "→ Định dạng chuẩn: DD/MM/YYYY (ví dụ: 15/03/2026)\n"
        "→ Hoặc: ngày 15 tháng 03 năm 2026\n"
        "→ KHÔNG viết thành câu, KHÔNG thêm mô tả.\n"
        "→ KHÔNG để trống hoặc viết '...' — phải có ngày cụ thể.\n"
        "\n"
        "✓ ĐÚNG: '15/03/2026', 'ngày 20 tháng 01 năm 2026'\n"
        "✗ SAI:  'trong năm 2026', 'gần đây', 'ngày...tháng...năm...'"
    ),

    "time_range": (
        "ĐÂY LÀ TRƯỜNG KHOẢNG THỜI GIAN / KỲ BÁO CÁO.\n"
        "→ Điền giai đoạn, quý, kỳ báo cáo từ tài liệu nguồn.\n"
        "→ Giữ đúng cách viết chuẩn hành chính:\n"
        "   'Quý I/2026', '6 tháng đầu năm 2026', 'giai đoạn 2021-2025'\n"
        "   'từ tháng 01/2026 đến tháng 06/2026', 'năm 2025'\n"
        "→ KHÔNG viết thành câu mô tả.\n"
        "\n"
        "✓ ĐÚNG: 'Quý I/2026', '6 tháng đầu năm 2026'\n"
        "✗ SAI:  'thời gian qua', 'giai đoạn vừa qua'"
    ),

    "number": (
        "ĐÂY LÀ TRƯỜNG SỐ LIỆU / CON SỐ CỤ THỂ.\n"
        "→ Tìm con số chính xác từ tài liệu nguồn.\n"
        "→ GIỮ NGUYÊN đơn vị: tỷ đồng, triệu người, nghìn lệnh/ngày, dự án...\n"
        "→ Dùng dấu phẩy phân cách hàng nghìn: 1.234,5 tỷ đồng\n"
        "→ Nếu không tìm được số → ghi: 'chưa có số liệu cụ thể trong tài liệu nguồn'\n"
        "→ KHÔNG bịa số, KHÔNG ước tính, KHÔNG làm tròn.\n"
        "\n"
        "✓ ĐÚNG: '1.847,3 tỷ đồng', '9,8 triệu KHCN', '156 dự án'\n"
        "✗ SAI:  'khoảng 2.000 tỷ', 'hàng nghìn', 'một lượng lớn'"
    ),

    "percentage": (
        "ĐÂY LÀ TRƯỜNG TỶ LỆ PHẦN TRĂM.\n"
        "→ Tìm % chính xác từ tài liệu nguồn.\n"
        "→ Giữ đúng số thập phân: 94,5% (không làm tròn thành 95%)\n"
        "→ Có thể kèm so sánh nếu source có: 'đạt 94,5%, tăng 12% so với cùng kỳ'\n"
        "→ KHÔNG bịa %, KHÔNG ước tính.\n"
        "\n"
        "✓ ĐÚNG: '94,5%', 'đạt 87,2%, tăng 5,3 điểm %'\n"
        "✗ SAI:  'tỷ lệ cao', 'gần 100%', 'phần lớn'"
    ),

    "name": (
        "ĐÂY LÀ TRƯỜNG TÊN / MÃ SỐ / VĂN BẢN.\n"
        "→ Điền chính xác: tên người, tên đơn vị, số văn bản, mã dự án.\n"
        "→ Viết hoa đúng: BIDV, NHNN, NQ57, Bộ Công an\n"
        "→ Số văn bản đúng format: số 123/QĐ-BIDV ngày 15/03/2026\n"
        "→ KHÔNG viết câu — chỉ điền tên/mã.\n"
        "\n"
        "✓ ĐÚNG: 'Nghị quyết số 57-NQ/TW', 'Ban Công nghệ Thông tin'\n"
        "✗ SAI:  'nghị quyết liên quan', 'đơn vị chức năng'"
    ),

    "short": (
        "ĐÂY LÀ TRƯỜNG NGẮN (cụm từ, nhãn, tiêu đề).\n"
        "→ Điền 1 cụm từ hoặc nhãn cụ thể, thường 3-25 ký tự.\n"
        "→ Không cần câu hoàn chỉnh.\n"
        "→ Lấy chính xác từ tài liệu nguồn hoặc ngữ cảnh xung quanh.\n"
        "\n"
        "✓ ĐÚNG: 'Hoàn thành', 'Đang triển khai', 'Ban CNTT'\n"
        "✗ SAI:  'Mục tiêu đã được hoàn thành trong quý I theo kế hoạch'"
    ),

    "sentence": (
        "ĐÂY LÀ TRƯỜNG 1-3 CÂU VĂN LIỀN MẠCH.\n"
        "→ Viết 1-3 câu hoàn chỉnh, mạch lạc, văn phong hành chính.\n"
        "→ Mỗi câu PHẢI chứa ít nhất 1 dữ kiện cụ thể (số, tên, ngày).\n"
        "→ Dùng câu trần thuật — KHÔNG dùng gạch đầu dòng.\n"
        "→ Văn phong: súc tích, khách quan, chủ ngữ rõ ràng.\n"
        "\n"
        "✓ ĐÚNG: 'BIDV hoàn thành triển khai hệ thống Core Banking mới với 156 chi nhánh kết nối trực tiếp, "
        "đạt 98% kế hoạch Quý I/2026.'\n"
        "✗ SAI:  'Đã triển khai tốt các nhiệm vụ được giao, kết quả đạt khả quan.'"
    ),

    "paragraph": (
        "ĐÂY LÀ TRƯỜNG ĐOẠN VĂN DÀI (nhiều câu, liền mạch).\n"
        "→ Viết đoạn văn hoàn chỉnh, nhiều câu, mỗi câu có dữ kiện cụ thể.\n"
        "→ CẤU TRÚC: câu mở đầu tổng quát → các câu chi tiết số liệu → câu kết.\n"
        "→ Văn phong LIỀN MẠCH — dùng từ nối: 'Trong đó,', 'Cụ thể,', 'Bên cạnh đó,'\n"
        "→ KHÔNG dùng gạch đầu dòng trừ khi placeholder gốc có gạch đầu dòng.\n"
        "→ Độ dài: TỐI THIỂU bằng placeholder gốc, nên dài hơn nếu có đủ dữ kiện.\n"
        "\n"
        "✓ ĐÚNG: 'Trong Quý I/2026, BIDV đã hoàn thành triển khai 3 dự án trọng điểm về chuyển đổi số. "
        "Cụ thể, dự án Core Banking thế hệ mới kết nối 156 chi nhánh với uptime 99,7%. "
        "Hệ thống thanh toán QR xử lý trung bình 2,3 triệu lệnh/ngày, tăng 45% so với Quý IV/2025. "
        "Nền tảng eKYC tích hợp VNeID hoàn thành xác thực 9,8 triệu KHCN.'\n"
        "✗ SAI:  'BIDV đã triển khai nhiều dự án quan trọng và đạt kết quả tích cực trong công tác chuyển đổi số.'"
    ),

    "bullet_list": (
        "ĐÂY LÀ TRƯỜNG DANH SÁCH / GẠH ĐẦU DÒNG.\n"
        "→ Viết theo format gạch đầu dòng, mỗi dòng là 1 ý riêng biệt.\n"
        "→ Dùng ký hiệu '- ' đầu mỗi dòng (hoặc giữ đúng format placeholder gốc).\n"
        "→ Mỗi gạch đầu dòng PHẢI có ít nhất 1 dữ kiện cụ thể.\n"
        "→ Các dòng phân cách bằng ký tự xuống dòng '\\n'.\n"
        "→ Thường dùng cho: liệt kê kết quả, danh sách dự án, các bước triển khai.\n"
        "\n"
        "✓ ĐÚNG:\n"
        "'- Dự án Core Banking: kết nối 156 chi nhánh, uptime 99,7%\\n"
        "- Hệ thống QR Pay: 2,3 triệu lệnh/ngày, tăng 45%\\n"
        "- eKYC VNeID: xác thực 9,8 triệu KHCN'\n"
        "✗ SAI:\n"
        "'- Đã triển khai nhiều dự án\\n- Kết quả tốt\\n- Tiếp tục phát huy'"
    ),
}


# ═══════════════════════════════════════════════════════════════════════════════
# IMPROVED: Field Classification
# ═══════════════════════════════════════════════════════════════════════════════

def _classify_field(placeholder: str, context: str) -> str:
    """
    Phân loại field thông minh hơn:
    1. Check bullet/list structure trước (ưu tiên cao nhất cho format)
    2. Check date/time_range
    3. Check percentage (trước number vì % cũng chứa số)
    4. Check number
    5. Check name/entity
    6. Phân loại theo độ dài: short / sentence / paragraph
    """
    ph     = placeholder.strip()
    ctx    = context.strip()
    ph_len = len(ph)

    # ═══ 1. BULLET LIST — check cấu trúc placeholder có gạch đầu dòng không ═══
    bullet_score = 0
    for pat in _BULLET_PATTERNS:
        if pat.search(ph):
            bullet_score += 1
    # Nhiều dòng + có pattern list → chắc chắn là bullet
    line_count = ph.count('\n') + 1
    if bullet_score >= 1 and line_count >= 2:
        return "bullet_list"
    # Placeholder dài + nhiều dấu ; hoặc nhiều dòng → likely list
    if ph_len > 100 and (ph.count(';') >= 2 or line_count >= 3):
        return "bullet_list"
    # Placeholder có nhiều "-" đầu dòng
    if len(_re.findall(r'^\s*[-–—•]', ph, _re.M)) >= 2:
        return "bullet_list"

    # ═══ 2. DATE — ngày cụ thể ═══════════════════════════════════════════════
    for pat in _DATE_PATTERNS:
        if pat.search(ph):
            # Nếu placeholder dài > 60 → không phải field ngày thuần túy
            if ph_len <= 60:
                return "date"

    # Placeholder chỉ toàn dấu lửng/gạch → xem context
    if _re.fullmatch(r'[.\u2026\s/\-–—…]+', ph):
        _FUZZY_DATE = _re.compile(
            r'\b(ngày|tháng|năm|quý|kỳ|thời\s*gian|dd|mm|yyyy)\b', _re.I
        )
        if _FUZZY_DATE.search(ctx):
            return "date"

    # Placeholder có từ khóa thời gian nhưng ngắn
    _FUZZY_DATE2 = _re.compile(
        r'\b(ngày|tháng|năm)\b', _re.I
    )
    if _FUZZY_DATE2.search(ph) and ph_len <= 40:
        return "date"

    # ═══ 3. TIME RANGE — khoảng thời gian / kỳ báo cáo ══════════════════════
    for pat in _TIME_RANGE_PATTERNS:
        if pat.search(ph):
            if ph_len <= 80:
                return "time_range"

    # Context có keyword kỳ báo cáo + placeholder ngắn
    _TIME_CTX = _re.compile(
        r'\b(kỳ báo cáo|giai đoạn|quý|6 tháng|cả năm|thời kỳ)\b', _re.I
    )
    if _TIME_CTX.search(ctx) and ph_len <= 50 and _re.fullmatch(r'[.\u2026\s/\-–—…\d]+', ph):
        return "time_range"

    # ═══ 4. PERCENTAGE — tỷ lệ % ═════════════════════════════════════════════
    pct_score = 0
    for pat in _PERCENTAGE_PATTERNS:
        if pat.search(ph):
            pct_score += 1
    if pct_score >= 1 and ph_len <= 40:
        return "percentage"

    # ═══ 5. NUMBER — con số cụ thể ═══════════════════════════════════════════
    num_score = 0
    for pat in _NUMBER_PATTERNS:
        if pat.search(ph):
            num_score += 1
    # Placeholder chỉ là "XXX" hoặc "..." + context có đơn vị
    if _re.fullmatch(r'[.\u2026Xx\s]+', ph) and ph_len <= 20:
        _NUM_CTX = _re.compile(r'\b(tỷ|triệu|nghìn|%|đồng|lệnh|người|dự án)\b', _re.I)
        if _NUM_CTX.search(ctx):
            return "number"
    if num_score >= 1 and ph_len <= 60:
        return "number"

    # ═══ 6. NAME / ENTITY ════════════════════════════════════════════════════
    name_score = 0
    for pat in _NAME_PATTERNS:
        if pat.search(ph):
            name_score += 1
    # Context có "tên", "số văn bản", "đơn vị" → name
    _NAME_CTX = _re.compile(
        r'\b(tên|họ và tên|chức danh|số văn bản|đơn vị|cơ quan|bên)\b', _re.I
    )
    if _NAME_CTX.search(ctx) and ph_len <= 80:
        return "name"
    if name_score >= 2 and ph_len <= 60:
        return "name"

    # ═══ 7. PHÂN LOẠI THEO ĐỘ DÀI ═══════════════════════════════════════════
    if ph_len <= 25:
        return "short"
    elif ph_len <= 120:
        return "sentence"
    elif ph_len <= 300:
        # Check lại — có thể là bullet list ngắn
        if ph.count(';') >= 2 or len(_re.findall(r'[-–•]', ph)) >= 2:
            return "bullet_list"
        return "sentence"
    else:
        # > 300 chars
        # Check bullet list cho đoạn dài
        if bullet_score >= 1 or ph.count(';') >= 3 or line_count >= 3:
            return "bullet_list"
        return "paragraph"


def _parse_highlights(doc_xml_path: str) -> List[Dict]:
    raw = open(doc_xml_path, "rb").read()
    tree = ET.fromstring(raw)
    fields = []

    all_paras = list(tree.iter(f"{{{NS_W}}}p"))

    for para_idx, para in enumerate(all_paras):
        runs = list(para.iter(f"{{{NS_W}}}r"))
        full_text = "".join(
            (r.find(f"{{{NS_W}}}t").text or "")
            for r in runs if r.find(f"{{{NS_W}}}t") is not None
        )

        # Thu thập highlights — vàng = replace, bất kỳ màu khác = insert_below
        yellow_texts = []
        green_texts  = []
        for run in runs:
            rpr = run.find(f"{{{NS_W}}}rPr")
            t   = run.find(f"{{{NS_W}}}t")
            if rpr is not None and t is not None and t.text:
                hl  = rpr.find(f"{{{NS_W}}}highlight")
                shd = rpr.find(f"{{{NS_W}}}shd")
                shd_val = shd.get(f"{{{NS_W}}}fill") if shd is not None else None
                val = hl.get(f"{{{NS_W}}}val") if hl is not None else None
                if val == "yellow" or shd_val == "FFFF00":   # highlight vàng → replaces
                    yellow_texts.append(t.text)
                elif val is not None:   # bất kỳ màu nào khác vàng → insert_below
                    green_texts.append(t.text)

        highlights = yellow_texts or green_texts
        if not highlights:
            continue

        # field_mode: vàng = replace, xanh = insert_below
        field_mode  = "replace" if yellow_texts else "insert_below"
        placeholder = "".join(highlights)

        # Context: đoạn trước + hiện tại + sau (mở rộng hơn — 2 đoạn trước/sau)
        prev_texts = []
        for offset in [1, 2]:
            if para_idx >= offset:
                prev_runs = list(all_paras[para_idx-offset].iter(f"{{{NS_W}}}r"))
                prev_t = "".join(
                    (r.find(f"{{{NS_W}}}t").text or "")
                    for r in prev_runs if r.find(f"{{{NS_W}}}t") is not None
                )
                if prev_t.strip():
                    prev_texts.insert(0, prev_t[:120])

        next_texts = []
        for offset in [1, 2]:
            if para_idx + offset < len(all_paras):
                next_runs = list(all_paras[para_idx+offset].iter(f"{{{NS_W}}}r"))
                next_t = "".join(
                    (r.find(f"{{{NS_W}}}t").text or "")
                    for r in next_runs if r.find(f"{{{NS_W}}}t") is not None
                )
                if next_t.strip():
                    next_texts.append(next_t[:120])

        rich_context = " | ".join(
            prev_texts + [full_text[:250]] + next_texts
        ).strip()

        field_type = _classify_field(placeholder, rich_context)

        fields.append({
            "key":          f"para_{para_idx}",
            "para_idx":     para_idx,
            "placeholder":  placeholder,
            "context":      rich_context[:500],  # tăng từ 400 → 500
            "field_type":   field_type,
            "type_hint":    _FIELD_TYPE_HINTS.get(field_type, _FIELD_TYPE_HINTS["sentence"]),
            "field_mode":   field_mode,
        })

    return fields


# ─── Fill and export ──────────────────────────────────────────────────────────

def fill_and_export(
    docx_bytes:   bytes,
    field_values: List[Dict],
    apply_colors: bool = False,
) -> bytes:
    with tempfile.TemporaryDirectory() as tmp:
        docx_path  = os.path.join(tmp, "template.docx")
        unpack_dir = os.path.join(tmp, "unpacked")
        out_path   = os.path.join(tmp, "output.docx")

        with open(docx_path, "wb") as f:
            f.write(docx_bytes)
        with zipfile.ZipFile(docx_path, "r") as z:
            z.extractall(unpack_dir)

        doc_xml_path = os.path.join(unpack_dir, "word", "document.xml")
        raw = open(doc_xml_path, "rb").read()
        raw = _process_fields(raw, field_values, keep_highlight=apply_colors)
        open(doc_xml_path, "wb").write(raw)

        with zipfile.ZipFile(docx_path, "r") as orig:
            with zipfile.ZipFile(out_path, "w") as out_zip:
                for item in orig.infolist():
                    modified = os.path.join(unpack_dir, item.filename)
                    if os.path.isfile(modified):
                        out_zip.write(modified, item.filename,
                                      compress_type=item.compress_type)
                    else:
                        out_zip.writestr(item, orig.read(item.filename))

        return open(out_path, "rb").read()


def _build_new_para(value: str, ref_rpr: ET.Element) -> ET.Element:
    """Tạo paragraph mới với text và copy rPr từ run tham chiếu (bỏ highlight)."""
    import copy
    new_p = ET.Element(f"{{{NS_W}}}p")
    for i, line in enumerate(value.split('\n')):
        if i > 0:
            br_r = ET.SubElement(new_p, f"{{{NS_W}}}r")
            ET.SubElement(br_r, f"{{{NS_W}}}br")
        new_r  = ET.SubElement(new_p, f"{{{NS_W}}}r")
        if ref_rpr is not None:
            new_rpr = copy.deepcopy(ref_rpr)
            # Xóa highlight khỏi rPr mới
            hl = new_rpr.find(f"{{{NS_W}}}highlight")
            if hl is not None:
                new_rpr.remove(hl)
            new_r.append(new_rpr)
        t_el = ET.SubElement(new_r, f"{{{NS_W}}}t")
        t_el.text = line
        if line.startswith(' ') or line.endswith(' '):
            t_el.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')
    return new_p


def _find_parent(tree: ET.Element, child: ET.Element):
    """Tìm parent element của child."""
    for parent in tree.iter():
        if child in list(parent):
            return parent
    return None


def _process_fields(raw: bytes, field_values: List[Dict], keep_highlight: bool = False) -> bytes:
    """
    Xử lý fields theo field_mode:
    - replace      (🟡 vàng): ghi đè text tại chỗ
    - insert_below (🟢 xanh): giữ nguyên dòng tiêu đề, chèn paragraph mới bên dưới
    """
    val_map  = {fv["field_key"]: fv.get("final_value", "") for fv in field_values}
    mode_map = {fv["field_key"]: fv.get("field_mode", "replace") for fv in field_values}

    xml_str       = raw.decode("utf-8")
    root_open_end = xml_str.find('>') + 1
    root_tag_end  = xml_str.find('>', root_open_end) + 1
    orig_hdr      = xml_str[:root_tag_end]

    tree  = ET.fromstring(raw)
    paras = list(tree.iter(f"{{{NS_W}}}p"))

    inserts = []

    for para_idx, para in enumerate(paras):
        key   = f"para_{para_idx}"
        value = val_map.get(key)
        if value is None:
            continue

        mode = mode_map.get(key, "replace")

        h_runs = []
        for run in para.iter(f"{{{NS_W}}}r"):
            rpr = run.find(f"{{{NS_W}}}rPr")
            t   = run.find(f"{{{NS_W}}}t")
            if rpr is not None and t is not None and t.text:
                hl     = rpr.find(f"{{{NS_W}}}highlight")
                val_hl = hl.get(f"{{{NS_W}}}val") if hl is not None else None
                if val_hl is not None:
                    h_runs.append((t, rpr, hl))

        if not h_runs:
            continue

        first_t, first_rpr, first_hl = h_runs[0]

        if mode == "replace":
            first_t.text = value
            if value.startswith(' ') or value.endswith(' '):
                first_t.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')

            if not keep_highlight:
                first_rpr.remove(first_hl)
                for t, rpr, hl in h_runs[1:]:
                    t.text = ""
                    rpr.remove(hl)
            else:
                for t, rpr, hl in h_runs[1:]:
                    t.text = ""

        elif mode == "insert_below":
            for t, rpr, hl in h_runs:
                rpr.remove(hl)

            new_p  = _build_new_para(value, first_rpr)
            parent = _find_parent(tree, para)
            if parent is not None:
                inserts.append((parent, para, new_p))

    for parent, ref_para, new_p in inserts:
        children = list(parent)
        for i, child in enumerate(children):
            if child is ref_para:
                parent.insert(i + 1, new_p)
                break

    new_xml = ET.tostring(tree, encoding="unicode", xml_declaration=False)
    new_end = new_xml.find('>') + 1
    new_xml = orig_hdr + new_xml[new_end:]

    return new_xml.encode("utf-8")


def _replace_highlights_raw(raw: bytes, field_values: List[Dict], apply_colors: bool) -> bytes:
    return _process_fields(raw, field_values)


# ─── Extract text from source files ──────────────────────────────────────────

def extract_text_from_file(file_bytes: bytes, filename: str) -> str:
    ext = os.path.splitext(filename)[1].lower()
    with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name
    try:
        if ext == ".pdf":
            r = subprocess.run(["pdftotext", tmp_path, "-"], capture_output=True)
            return r.stdout.decode("utf-8", errors="replace")[:60000]
        elif ext in (".docx", ".doc"):
            r = subprocess.run(
                ["pandoc", "--track-changes=all", tmp_path, "-t", "plain"],
                capture_output=True
            )
            return r.stdout.decode("utf-8", errors="replace")[:60000]
        elif ext in (".xlsx", ".xls"):
            return _extract_text_from_excel(file_bytes, ext)
        else:
            with open(tmp_path, errors="replace") as f:
                return f.read()[:60000]
    finally:
        os.unlink(tmp_path)

def _extract_text_from_excel(file_bytes: bytes, ext: str) -> str:
    """
    Đọc Excel → text dạng bảng, mỗi sheet là 1 section.
    """
    import io
    lines = []
    try:
        if ext == ".xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                lines.append(f"=== Sheet: {sheet_name} ===")
                for row in ws.iter_rows():
                    cells = []
                    for cell in row:
                        val = cell.value
                        if val is not None:
                            cells.append(str(val).strip())
                    if any(cells):
                        lines.append(" | ".join(cells))
        elif ext == ".xls":
            import xlrd
            wb = xlrd.open_workbook(file_contents=file_bytes)
            for sheet_name in wb.sheet_names():
                ws = wb.sheet_by_name(sheet_name)
                lines.append(f"=== Sheet: {sheet_name} ===")
                for i in range(ws.nrows):
                    cells = [str(ws.cell_value(i, j)).strip()
                             for j in range(ws.ncols)]
                    if any(cells):
                        lines.append(" | ".join(cells))
    except Exception as e:
        return f"(Lỗi đọc Excel: {e})"

    return "\n".join(lines)[:60000]